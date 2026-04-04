import os
import io
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
import csv


# 标准字段定义
STANDARD_FIELDS = {
    "patient_name": "患者姓名",
    "record_no": "病案号",
    "gender": "性别",
    "age": "年龄",
    "diagnosis": "诊断",
    "phone": "电话",
    "address": "住址",
    "discharge_department": "出院科室",
    "admission_date": "入院日期",
    "discharge_date": "出院日期",
    "attending_doctor": "主管医师",
}

FIELD_ALIASES = {
    "patient_name": ["患者姓名", "姓名"],
    "record_no": ["病案号", "住院号"],
    "gender": ["性别"],
    "age": ["年龄"],
    "diagnosis": ["诊断", "出院诊断", "门诊诊断", "入院诊断"],
    "phone": ["电话", "联系电话", "联系方式", "手机号", "手机号码"],
    "address": ["住址", "家庭住址", "现住址", "联系地址", "地址", " 现住址"],
    "discharge_department": ["出院科室", "科室"],
    "admission_date": ["入院日期"],
    "discharge_date": ["出院日期"],
    "attending_doctor": ["主管医师", "主治医师", "住院医师"],
}


def generate_import_template():
    """生成标准导入模板Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "出院患者"
    headers = list(STANDARD_FIELDS.values())
    ws.append(headers)
    # 示例数据
    ws.append(["张三", "202601001", "男", 45, "高血压", "13800138000", "平顶山市湛河区示例路1号", "内分泌科", "2026-01-01", "2026-01-10", "王医师"])

    from flask import current_app
    template_dir = current_app.config.get("UPLOAD_FOLDER", "uploads")
    os.makedirs(template_dir, exist_ok=True)
    filepath = os.path.join(template_dir, "出院患者导入模板.xlsx")
    wb.save(filepath)
    return filepath


def preview_columns(file_storage):
    """预览文件列名和前5行数据"""
    filename = file_storage.filename.lower()
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return _preview_excel(file_storage)
    elif filename.endswith(".csv"):
        return _preview_csv(file_storage)
    else:
        raise ValueError("不支持的文件格式，请上传Excel或CSV文件")


def _preview_excel(file_storage):
    wb = load_workbook(file_storage, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_row=6, values_only=True))
    if not rows:
        raise ValueError("文件为空")
    columns = [str(c) if c else f"列{i+1}" for i, c in enumerate(rows[0])]
    sample = []
    for row in rows[1:6]:
        sample.append([_cell_to_str(c) for c in row])
    return columns, sample


def _preview_csv(file_storage):
    content = file_storage.read().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    rows = []
    for i, row in enumerate(reader):
        if i >= 6:
            break
        rows.append(row)
    if not rows:
        raise ValueError("文件为空")
    columns = rows[0]
    sample = rows[1:6]
    return columns, sample


def _cell_to_str(cell):
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    return str(cell)


def parse_patient_file(file_storage, column_map=None):
    """解析Excel/CSV文件，返回患者字典列表"""
    filename = file_storage.filename.lower()
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return _parse_excel(file_storage, column_map)
    elif filename.endswith(".csv"):
        return _parse_csv(file_storage, column_map)
    else:
        raise ValueError("不支持的文件格式")


def _build_field_map(headers, column_map):
    """根据列映射或自动匹配，建立列索引到字段名的映射"""
    field_map = {}
    normalized_headers = {str(h).strip(): i for i, h in enumerate(headers) if h}
    if column_map:
        # column_map: {"patient_name": "患者姓名", "record_no": "住院号", ...}
        for field, col_name in column_map.items():
            col_name = str(col_name).strip()
            if col_name in normalized_headers:
                field_map[normalized_headers[col_name]] = field
    for field, aliases in FIELD_ALIASES.items():
        if field in field_map.values():
            continue
        for alias in aliases:
            idx = normalized_headers.get(alias.strip())
            if idx is not None:
                field_map[idx] = field
                break
    return field_map


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    val = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(val):
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _parse_age(val):
    if val is None:
        return None
    text = str(val).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        try:
            return int(digits)
        except ValueError:
            return None
    return _parse_int(val)


def _row_to_patient(row, field_map):
    p = {}
    for idx, field in field_map.items():
        if idx < len(row):
            val = row[idx]
            if field in ("admission_date", "discharge_date"):
                p[field] = _parse_date(val)
            elif field == "age":
                p[field] = _parse_age(val)
            else:
                p[field] = str(val).strip() if val else ""
    return p


def _parse_excel(file_storage, column_map):
    wb = load_workbook(file_storage, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("文件数据为空（至少需要表头和一行数据）")
    headers = [str(c) if c else "" for c in rows[0]]
    field_map = _build_field_map(headers, column_map)
    if not field_map:
        raise ValueError("无法匹配任何字段，请检查列映射或使用标准模板")
    patients = []
    for row in rows[1:]:
        p = _row_to_patient(row, field_map)
        if p.get("patient_name"):
            patients.append(p)
    return patients


def _parse_csv(file_storage, column_map):
    content = file_storage.read().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    if len(rows) < 2:
        raise ValueError("文件数据为空")
    headers = rows[0]
    field_map = _build_field_map(headers, column_map)
    if not field_map:
        raise ValueError("无法匹配任何字段，请检查列映射或使用标准模板")
    patients = []
    for row in rows[1:]:
        p = _row_to_patient(row, field_map)
        if p.get("patient_name"):
            patients.append(p)
    return patients
